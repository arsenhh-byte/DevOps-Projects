""" using the excel sheet, I am going to do the following:
1.List each company with respective product count
2. List products with inventory less than 10
3. List each company with respective total inventory value
4. Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet (inventory* price)
Save the file programmatically
"""
import openpyxl
path = "/Users/goldenchild/Desktop/DevOps Projects/Automation with Python - Spreadsheet/inventory.xlsx"

# To open the workbook
inv_file = openpyxl.load_workbook(path)
product_list = inv_file["Sheet1"]

# Task uno: How many products we are able to calculate per supplier and list supplier with the respective number of products in form of a dictionary 
product_per_supplier = {}
total_value_per_supplier = {}
products_invetory_under10 = {}

# Getting the info from the excel sheet
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory_value = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    # Getting the products a supplier has and putting it into a dictionary. If the products already exist, we increment them
    #Task 1 done...Listing each company with respective product count
 
    if supplier_name in product_per_supplier:
        current_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_products + 1
    else:
        product_per_supplier[supplier_name] = 1


#2. List products with inventory less than 10 
    if inventory_value < 10:
        products_invetory_under10[int(product_num)] = int(inventory_value)




#  3. List each company with respective total inventory value
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] =  current_value + (price * inventory_value)
    
    else:
        total_value_per_supplier[supplier_name] = price * inventory_value
       
    
# 4. Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet (inventory * price)
    inventory_price.value = inventory_value * price




inv_file.save("Inventory_with_total_value.xlsx")

print(f"Exercise 1 done {product_per_supplier}")
print(f" Exercise 2 done {total_value_per_supplier}")
print(f" Exercise 3 done {products_invetory_under10}")